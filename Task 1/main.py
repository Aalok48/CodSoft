from tkinter import *
from openpyxl import Workbook, load_workbook
from tkcalendar import Calendar, DateEntry
from tkinter import messagebox as tmsg, ttk
import pandas as pd

root = Tk()                                                     # creating the root for tkinter
root.geometry("375x500")                                        # defining the geometry for the tkinter window
root.resizable(False, False)                                    # making the tkinter window not resizable
root.title("To-Do List")                                        # title for the tkinter
root.iconbitmap("D:\CodSoft\Task 1\photos\download1.ico")       # specifying the icon for the tkinter

# welcome frame for the welcome text
welcome_label_frame = Frame(root, width=375, height=60)
welcome_label_frame.pack(pady=35)

# lable for the welcome frame 
welcome_label = Label(welcome_label_frame, text="Welcome to TO-DO List", font=35)
welcome_label.pack()

# main frame that contains the create, update and track options
options_frame = Frame(root, width=375, height=275)
options_frame.place(x=0, y=125)

# function part
# this is create funtion
def create():
    # this clears the existing widgets and buttons from the main page
    for widget in root.winfo_children():
        widget.destroy()

    root.geometry('700x550')

    # welcome frame for the add list
    create_welcome_frame = Frame(root, width=375, height=75)
    create_welcome_frame.pack(pady=35)

    # text lable for the add list
    create_welcome_text = Label(create_welcome_frame, text='Add your to-do items', font=45)
    create_welcome_text.place(x=85,y=35)

    # frame where all the text options and save button
    create_menu_frame = Frame(root, width=475, height=375)
    create_menu_frame.pack(pady=0)

    # area for the set_to_do
    set_to_do = Label(create_menu_frame, width=12, font=10, text='To-Do activity: ')
    set_to_do.place(x=0, y=30)

    # main text area for the add list
    entry1 = Text(create_menu_frame, width=25, height=5, borderwidth=2, font=10, relief=GROOVE)
    entry1.place(x=135, y=0)

    # lable for the category
    set_category = Label(create_menu_frame, width=12, font=10, text='Category: ')
    set_category.place(x=-18, y=165)

    # relatively smaller text area for the category
    category = Entry(create_menu_frame, width=10, font=10, borderwidth=2, relief=GROOVE)
    category.place(x=135, y=165)

    # lable for the date and time
    date_time_lable = Label(create_menu_frame, font=10, text="Choose Date: ")
    date_time_lable.place(x=0, y=225)

    def show_calender():
        top = Toplevel(root)
        cal = Calendar(top, selectmode = 'day', year = 2023, month = 1, day = 1)
        cal.pack()

        def grab_date():
            my_date = cal.get_date()
            date_entry.delete(0, END)
            date_entry.insert(0, my_date)
            top.destroy()

        Button(top, text='select', command=grab_date).pack()

    date_entry = DateEntry(create_menu_frame, width=12, background = 'darkblue', foreground = 'white', borderwidth = 2)
    date_entry.place(x=135, y=230)

    # function is called when the save button is clicked
    def save_items():
        to_do_entry = entry1.get('1.0', END)
        to_do_category = category.get()
        to_do_date = date_entry.get()

        if to_do_entry.strip() == "" or to_do_category == "":
            tmsg.showerror('Error', 'Text or category cannot be empty')
        else:
            try:
            # Try to load the existing workbook
                wb = load_workbook('to_do_list.xlsx')
                sheet = wb.active
            except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
                wb = Workbook()
                sheet = wb.active
                sheet.title = "To-Do List"
                sheet['A1'] = 'Task'
                sheet['B1'] = 'Category'
                sheet['C1'] = 'is_todo'
                sheet['D1'] = 'Date'

        # Get the last row number
            last_row = sheet.max_row

        # Write data to the next available row
            sheet.cell(row=last_row + 1, column=1).value = to_do_entry
            sheet.cell(row=last_row + 1, column=2).value = to_do_category
            sheet.cell(row=last_row + 1, column=3).value = True
            sheet.cell(row=last_row + 1, column=4).value = to_do_date


            wb.save('to_do_list.xlsx')   
            entry1.delete('1.0', END)
            category.delete(0, END)

    save_button = Button(create_menu_frame, text='Save', command=save_items, width=15, relief=GROOVE)
    save_button.place(x = 150, y=300)
    

# this is update function
def update():
    for widget in root.winfo_children():
        widget.destroy()

    root.geometry('603x230')

    df = pd.read_excel('to_do_list.xlsx')
    

    # Filter rows where 'is_todo' is True
    todo_rows = df[df['is_todo'] == True]

    tree = ttk.Treeview(root, columns=("Task", "Category", "Date"), show="headings")
    tree.grid(row=0, column=0, sticky='nsew')

    tree.heading("Task", text="Task")
    tree.heading("Category", text="Category")
    tree.heading("Date", text="Date")

    for index, row in todo_rows.iterrows():
       tree.insert("", END, values=(row['Task'], row['Category'], row['Date']))


    # text = Text(root)
    # text.grid(row = 0, column = 0, sticky='nsew')
    # scrollbar = Scrollbar(root, orient="vertical", command=text.yview, width=20)
    # scrollbar.grid(row=0, column=1, sticky='ns')
    # text.config(yscrollcommand=scrollbar.set)

    # for index, row in todo_rows.iterrows():
    #     text.insert(END, f"Row {index+1}:\n")
    #     for col in row.index:
    #         text.insert(END, f"{col}: {row[col]}\n")
    #     text.insert(END, "\n")

    # Create a scrollbar for the Treeview
    scrollbar = Scrollbar(root, orient="vertical", command=tree.yview)
    scrollbar.grid(row=0, column=1, sticky='ns')
    tree.config(yscrollcommand=scrollbar.set)

# Example usage:
    excel_file = "to_do_list.xlsx"  # Replace with your Excel file path
    update(excel_file)
    

# this is track function
def track():
    print('Track button clicks')

# three button namingly create, update and track button are made
# this is create button
button1 = Button(options_frame, text='Create', width=15, borderwidth=1, command=create, font=1)
button1.place(x=105, y=45)

# this is update button
button2 = Button(options_frame, text='Update', width=15, borderwidth=1, command=update, font=1)
button2.place(x=105, y=115)

# this is track button
button3 = Button(options_frame, text='Track', width=15, borderwidth=1, command=track, font=1)
button3.place(x=105, y=185)

root.mainloop()